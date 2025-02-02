from flask import Flask, request
from linebot import LineBotApi, WebhookHandler
from linebot.models import (MessageEvent,
                            TextMessage,
                            TextSendMessage)

from openpyxl import load_workbook

channel_secret = "bd64f7ffb48f1aada1527ee852c4fda0"
channel_access_token = "RrewYzqHmm7KSmeVPPowpWx44BrF1ABUhMxaFvb2wwPlLf/Ct6M0w+mQpZBOkTt5RxwtW0jUmT97K1JTKn7vW968Qdgo3btfXw425HYFsaanXy/YcqXSRMePK8r4pdCi6b6GkoSNfTQz8guUo69iuAdB04t89/1O/w1cDnyilFU="

line_bot_api = LineBotApi(channel_access_token)
handler = WebhookHandler(channel_secret)

app = Flask(__name__)

@app.route("/", methods=["GET","POST"])
def home():
    try:
        signature = request.headers["X-Line-Signature"]
        body = request.get_data(as_text=True)
        handler.handle(body, signature)
    except:
        pass
    
    return "Hello Line Chatbot"

@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    text = event.message.text
    print(text)

    wb = load_workbook(filename='product.xlsx', read_only=True)
    ws = wb['Sheet1']

    if (text.startswith("ค้นหา")): # "ค้นหา TV7788"
        text_id = text.strip("ค้นหา").replace(" ", "") # "TV7788"
        print(text_id) # "TV7788"
        
    for i in range(2,ws.max_row+1):              # เริ่มอ่านที่ cell A2 เป็นต้นไป
        if text_id == str(ws["A"+str(i)].value): # ค้นหารหัสสินค้า
            product_name = str(ws["B"+str(i)].value) # ชื่อสินค้า
            product_price = str(ws["C"+str(i)].value) # ราคา
            text_out = "รหัสสินค้า " + text_id + " คือ " + product_name + " ราคา " + product_price + " บาท"
            break
        else:
            text_out = "ไม่พบรหัสสินค้านี้"
    wb.close()
    
    line_bot_api.reply_message(event.reply_token,
                             TextSendMessage(text=text_out))

if __name__ == "__main__":          
    app.run()

